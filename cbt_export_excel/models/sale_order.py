import base64
import os
import xlwt
from io import BytesIO, StringIO

import pandas as pd
from openpyxl.workbook import Workbook
import xlsxwriter

from odoo import api, fields, models


class SaleOrder(models.Model):
    _inherit = "sale.order"

    def print_xlsx_report(self):
        return {
            'type': 'ir.actions.act_url',
            'url': '/invoicing/excel_report/%s' % (self.id),
            'target': 'new',
        }

    def get_report_lines(self):
        invoice_list = []
        # for move in self.env['account.move'].search([]):
        for order in self.env['sale.order'].search([('id', '=', self.id)]).order_line:
            # results = self.env['sale.order.line'].search([('move_id', '=', move.id)])
            # if results:
            if order:
                print('order display_type---------->', order.display_type)
                line = {
                    'd_type': order.display_type,
                    'description': order.name,
                    'qty': order.product_uom_qty,
                    'unit_price': order.price_unit,
                    'total_price': order.price_subtotal,
                    'estimation_cost': order.estimate_cost,
                    # 'move_name': move.name,
                    # 'partner_id': move.partner_id.name,
                }
                invoice_list.append(line)
        return invoice_list

    #
    # def generate_excel_report(self):
    #     filename = 'filename.xlsx'
    #     workbook = xlwt.Workbook(encoding="UTF-8")
    #     worksheet = workbook.add_sheet('Sheet 1')
    #     style = xlwt.easyxf('font: bold True, name Arial;')
    #     worksheet.write_merge(0, 1, 0, 3, 'your data that you want to show into excelsheet', style)
    #     fp = StringIO()
    #     workbook.save(fp)
    #     record_id = self.env['wizard.excel.report'].create({'excel_file': base64.encodestring(fp.getvalue()),
    #                                                         'file_name': filename}, )
    #     fp.close()
    #     return {'view_mode': 'form',
    #             'res_id': record_id,
    #             'res_model': 'wizard.excel.report',
    #             'view_type': 'form',
    #             'type': 'ir.actions.act_window',
    #             'context': context,
    #             'target': 'new',
    #             }
    #

    def action_export(self):
        data = [
            {
                'name': "Edan Stein",
                'phone': "1-411-426-8735",
                'email': "facilisis.magna@aol.couk",
                'address': "594-6075 Elementum Ave",
                'country': "Belgium"
            },
            {
                'name': "Gretchen Whitfield",
                'phone': "(323) 253-9734",
                'email': "ipsum@protonmail.net",
                'address': "Ap #783-9102 Augue. Rd.",
                'country': "Netherlands"
            },
            {
                'name': "Violet Brooks",
                'phone': "1-389-367-4883",
                'email': "montes.nascetur.ridiculus@outlook.edu",
                'address': "Ap #814-4695 Odio. Street",
                'country': "India"
            },
            {
                'name': "Ethan Espinoza",
                'phone': "(428) 503-8130",
                'email': "vestibulum.lorem@yahoo.couk",
                'address': "3416 Suspendisse Rd.",
                'country': "Russian Federation"
            },
            {
                'name': "Martin Dunlap",
                'phone': "1-341-689-0165",
                'email': "natoque.penatibus@aol.couk",
                'address': "Ap #593-870 Rhoncus. Ave",
                'country': "Poland"
            }
        ]

        workbook = xlsxwriter.Workbook("/home/oka/Desktop/AllAboutPythonExcel.xlsx")
        worksheet = workbook.add_worksheet("firstSheet")

        worksheet.write(0, 0, "#")
        worksheet.write(0, 1, "Name")
        worksheet.write(0, 2, "Phone")
        worksheet.write(0, 3, "Email")
        worksheet.write(0, 4, "Address")
        worksheet.write(0, 5, "Country")

        for index, entry in enumerate(data):
            worksheet.write(index + 1, 0, str(index))
            worksheet.write(index + 1, 1, entry["name"])
            worksheet.write(index + 1, 2, entry["phone"])
            worksheet.write(index + 1, 3, entry["email"])
            worksheet.write(index + 1, 4, entry["address"])
            worksheet.write(index + 1, 5, entry["country"])

        workbook.close()

    def execl_export(self):
        # dest_filename = 'FileName.xlsx'
        # fp = os.path.join(dest_filename)
        print('oooooooooooooooo', os.getcwd())
        workbook = xlsxwriter.Workbook('/home/oka/Desktop/Example3.xlsx')
        # workbook = xlsxwriter.Workbook(fp)
        worksheet = workbook.add_worksheet("My sheet")

        scores = (
            ['abc', 100],
            ['cde', 200],
            ['fghi', 565],
            ['jklmnop', 90],
        )
        row = 0
        col = 0
        for name, score in (scores):
            worksheet.write(row, col, name)
            worksheet.write(row, col + 1, score)
            row += 1
        workbook.close()

        desc_list = []
        qty_list = []
        price_list = []
        tot_price_list = []
        est_price_list = []
        out_list = []
        for rec in self.order_line:
            desc_list.append(rec.name)
            qty_list.append(rec.product_uom_qty)
            price_list.append(rec.price_unit)
            tot_price_list.append(rec.price_subtotal)
            est_price_list.append(rec.price_subtotal)
        # print(desc_list)
        # print(qty_list)
        # print(price_list)
        # print(tot_price_list)
        # print(est_price_list)
        # creating a data frame
        data = {
            # "Description": ["apple", "banana", "orange", "watermelon"],
            "Description": desc_list,
            # "Qty": [4, 2, 6, 3],
            "Qty": qty_list,
            # "Price": [120, 40, 80, 150],
            "Price": price_list,
            # "Total Price": [120, 40, 80, 150],
            "Total Price": tot_price_list,
            # "Estimation Cost": [120, 40, 80, 150]
            "Estimation Cost": est_price_list
        }

        df = pd.DataFrame(data)
        df.to_excel('/home/oka/Desktop/output_no_index.xlsx', index=False)
        # df.to_excel('./Desktop/test.xlsx', index=False)
        print(df)
